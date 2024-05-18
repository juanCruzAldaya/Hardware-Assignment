import os
import time
import openpyxl as xl
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.edge.options import Options as EdgeOptions
import logging


option = EdgeOptions()
option.add_argument("start-maximized")
driver = webdriver.Edge(options = option)

timeOut = 35
workbook = load_workbook(filename = "assignment_template.xlsx")
sheet = workbook.active




if sheet["A2"].value != None:
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=35, max_col=1):

        driver.get("https://support.accenture.com/now/nav/ui/classic/params/target/alm_hardware_list.do%3Fsysparm_query%3DGOTOasset_tagIN%26sysparm_first_row%3D1%26sysparm_view%3D%26sysparm_fixed_query%3Dinstall_status!%253D2")
        WebDriverWait(driver, timeOut).until(EC.title_is("Hardware | ServiceNow"))
        shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
        iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
        driver.switch_to.frame (iframe)
        for cell in row:
            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/span/div/div[1]/div/div[1]/a')))
                filterButton = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/span/div/div[1]/div/div[1]/a')
                filterButton.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/a')))
                serialButton = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[2]/div/a')
                serialButton.click()
            except TimeoutException:
                print ("Snow failed, run script again...")


            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div/input')))
                serialInput = driver.find_element(By.XPATH, '/html/body/div[5]/div/input')
                serialInput.send_keys("serial number")
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/ul/li/div')))
                serialPick = driver.find_element(By.XPATH, '/html/body/div[5]/ul/li/div')
                serialPick.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[4]/input')))
                searchBox = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[3]/td/table/tbody/tr/td[4]/input')
                searchBox.send_keys(cell.value)
            except TimeoutException:
                print ("Snow failed, run script again...")
            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[1]/button[1]')))
                runButton = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/span/div/div[4]/div/div/list_filter/div[1]/button[1]')
                runButton.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/span/div/div[7]/div[1]/table/tbody/tr/td[3]/a')))
                enterAsset = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/span/div/div[7]/div[1]/table/tbody/tr/td[3]/a')
                enterAsset.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select')))
                serialDropDown = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select')
                serialDropDown.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

            cellRow = str(cell.row)
            stateOption = sheet["c"+cellRow].value
            stateOptionToConc = str(stateOption)

            if  stateOption == 3:

                try:     
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')))
                    stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')
                    stateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")
                try:

                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select')))
                    subStateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select')
                    subStateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")
                time.sleep(1)
                try:
                    cellRow = str(cell.row)
                    subStateOption = sheet["e"+cellRow].value
                    subStateOptionToConc = str(subStateOption)
                    print(subStateOptionToConc)          
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select/option['+subStateOptionToConc+']')))
                    subStateReason = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select/option['+subStateOptionToConc+']')
                    subStateReason.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")

                try:
                    cellRow = str(cell.row)
                    userToAssign = sheet["b"+cellRow].value
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[1]/div[7]/div[2]/div[2]/input')))
                    userField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[1]/div[7]/div[2]/div[2]/input')
                    userField.send_keys(userToAssign)
                except TimeoutException:
                    print ("Snow failed, run script again...")

            elif stateOption == 7:
                try:
                    
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')))
                    stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')
                    stateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")
            else:
                try:
                    
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')))
                    stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[1]/div[2]/select/option['+stateOptionToConc+']')
                    stateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")

                try: 
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select')))
                    subStateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select')
                    subStateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")


                try:
                    cellRow = str(cell.row)
                    subStateOption = sheet["d"+cellRow].value
                    subStateOptionToConc = str(subStateOption)
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select/option['+subStateOptionToConc+']')))
                    subStateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select/option['+subStateOptionToConc+']')
                    subStateButton.click()

                except TimeoutException:
                    print ("Snow failed, run script again...")

                try:

                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[2]/div[2]/select')))
                    subStateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select')
                    subStateButton.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")
                time.sleep(2)    
                try:
                    cellRow = str(cell.row)
                    subStateOption = sheet["e"+cellRow].value
                    subStateOptionToConc = str(subStateOption)              
                    WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select/option['+subStateOptionToConc+']')))
                    subStateReason = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[3]/div[2]/select/option['+subStateOptionToConc+']')
                    subStateReason.click()
                except TimeoutException:
                    print ("Snow failed, run script again...")

            
            
            driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
            time.sleep(1)

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[17]/div[2]/input')))
                WKNArea = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[1]/div[2]/div[17]/div[2]/input')
                WKNArea.clear()
                cellRow = str(cell.row)
                WKNArea.send_keys(sheet["f"+cellRow].value)
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[2]/span/div/div[2]/div/div/div[2]/textarea')))
                commentArea = driver.find_element(By.XPATH, "/html/body/div[2]/form/span[2]/span/div/div[2]/div/div/div[2]/textarea")
                cellRow = str(cell.row)
                commentArea.send_keys(". " +sheet["g"+cellRow].value)
            except TimeoutException:
                print ("Snow failed, run script again...")

            try:
                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/span/span/nav/div/div[2]/span[1]/span[2]/span/button[2]')))
                saveButton = driver.find_element(By.XPATH, '/html/body/div[1]/span/span/nav/div/div[2]/span[1]/span[2]/span/button[2]  ')
                saveButton.click()
            except TimeoutException:
                print ("Snow failed, run script again...")

        driver.switch_to.new_window("tab")
else:
    print("EL ARCHIVO EXCEL NO TIENE ASSETS PARA AGREGAR")